import tkinter as tk
import pandas as pd
import spacy
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

nlp = spacy.load("en_core_web_sm")

def classify_incident(text):
    text = text.lower()
    accident_kw = {"accident", "crash", "collision", "ατύχημα", "σύγκρουση"}
    road_kw = {"road assistance", "breakdown", "tow", "οδική βοήθεια", "βλάβη"}

    tokens = set(text.split())
    if tokens & accident_kw:
        return "Accident Care"
    elif tokens & road_kw:
        return "Road Assistance"
    else:
        return "Other"

def save_to_excel(data):
    filename = "chat_answers.xlsx"
    df = pd.DataFrame([data])
    if os.path.exists(filename):
        with pd.ExcelWriter(filename, mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
            df.to_excel(writer, index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)
    else:
        df.to_excel(filename, index=False)

    wb = load_workbook(filename)
    ws = wb.active
    last_row = ws.max_row
    color_map = {
        "Accident Care": "FF9999",
        "Road Assistance": "CCFFCC",
        "Other": "FFFF99"
    }
    fill = PatternFill(start_color=color_map.get(data["Category"], "FFFFFF"),
                       end_color=color_map.get(data["Category"], "FFFFFF"),
                       fill_type="solid")
    for cell in ws[last_row]:
        cell.fill = fill
    wb.save(filename)
    print(f"Saved Excel file as: {filename}")

root = tk.Tk()
root.title("Chatbot")

chat_log = tk.Text(root, height=20, width=60)
chat_log.pack()

entry = tk.Entry(root, width=50)
send_btn = tk.Button(root, text="Send")

answers = {}
questions = [
    "Describe what happened.",
    "What is your plate registration number?",
    "What is your name?",
    "Where are you?",
    "Where would you like to go?"
]
columns = ["Incident Description", "RegNumber", "Name", "Location", "Destination"]
current_q = -1

def ask_next_question():
    global current_q
    current_q += 1
    if current_q < len(questions):
        chat_log.insert(tk.END, "Bot: " + questions[current_q] + "\n")
        if current_q == 0:
            # show entry and send for free text
            entry.pack(side=tk.LEFT)
            send_btn.pack(side=tk.LEFT)
        else:
            entry.pack(side=tk.LEFT)
            send_btn.pack(side=tk.LEFT)
    else:
        chat_log.insert(tk.END, "Bot: Thank you for the info! Goodbye.\n")
        save_to_excel(answers)
        entry.pack_forget()
        send_btn.pack_forget()

def on_send():
    global current_q
    user_text = entry.get().strip()
    if not user_text:
        return
    chat_log.insert(tk.END, "You: " + user_text + "\n")
    entry.delete(0, tk.END)

    if current_q == 0:
        answers[columns[0]] = user_text
        answers["Category"] = classify_incident(user_text)
        ask_next_question()
    else:
        answers[columns[current_q]] = user_text
        ask_next_question()

def on_yes():
    chat_log.insert(tk.END, "You: Yes\n")
    yes_btn.pack_forget()
    no_btn.pack_forget()
    ask_next_question()

def on_no():
    chat_log.insert(tk.END, "You: No\n")
    yes_btn.pack_forget()
    no_btn.pack_forget()
    chat_log.insert(tk.END, "Bot: Okay, have a nice day!\n")
    entry.pack_forget()
    send_btn.pack_forget()

def start():
    chat_log.insert(tk.END, "Bot: Hello, did something happen?\n")
    yes_btn.pack(side=tk.LEFT)
    no_btn.pack(side=tk.LEFT)

send_btn.config(command=on_send)
entry.bind("<Return>", lambda event: on_send())

yes_btn = tk.Button(root, text="Yes", command=on_yes)
no_btn = tk.Button(root, text="No", command=on_no)

entry.pack_forget()
send_btn.pack_forget()

start()
root.mainloop()
